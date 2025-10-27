from functools import partial, partialmethod
from typing import Any
from unittest import case
from django.db.models import Count, Sum, Case, When, IntegerField
from django.contrib import admin
from django.db.models.query import QuerySet
from django.forms import IntegerField
from django.http.request import HttpRequest
from . import models
from django.db.models.aggregates import Count, Max, Min, Sum
from django.db.models import F
from django.utils.html import format_html
from urllib.parse import urlencode
from django.urls import path, reverse
from django.db.models.functions import Concat
from django.db.models.fields import CharField
from django.db.models import F, Count, Sum, Value
from django.utils.translation import gettext_lazy as _
from django.db.models.functions import Concat
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
import csv
import json
import openpyxl  # 新增：用于处理Excel文件
from django.db import transaction
import io

# <font color="red">**新增点：标签过滤器**</font>
class TagFilter(admin.SimpleListFilter):
    """按标签过滤单词"""
    title = 'Tags'
    parameter_name = 'tag'

    def lookups(self, request, model_admin):
        # 根据用户权限返回不同的标签选项
        if request.user.is_superuser:
            tags = models.Tag.objects.all()
        else:
            tags = models.Tag.objects.filter(creator=request.user)
        return [(tag.id, tag.name) for tag in tags.distinct()]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(tags__id=self.value())
        return queryset

class EnglishWordMediaInline(admin.TabularInline):
    """单词多媒体文件内联管理"""
    model = models.EnglishWordMedia
    extra = 1
    fields = ['file', 'uploaded_at']
    readonly_fields = ['uploaded_at']
    can_delete = False  # 禁止删除已上传的附件

@admin.register(models.EnglishWord)
class EnglishWordAdmin(admin.ModelAdmin):
    """英文单词管理"""
    # <font color="red">**修改点：在列表显示中添加标签**</font>
    list_display = ['title', 'creator', 'created_at', 'media_count', 'get_tag_names']
    list_filter = ['created_at', TagFilter]  # <font color="red">**新增标签过滤器**</font>
    search_fields = ['title', 'explanation', 'notes', 'creator__username', 'creator__first_name', 'creator__last_name', 'tags__name']
    inlines = [EnglishWordMediaInline]
    
    # <font color="red">**修改点：在表单字段中添加标签**</font>
    filter_horizontal = ['tags']  # 水平多选标签
    
    actions = ['export_as_excel', 'export_as_csv', 'export_as_json']
    change_list_template = 'admin/englishword_change_list.html'
    
    # <font color="red">**修改点：排除创建者字段，但保留标签字段**</font>
    exclude = ['creator']
    
    def get_urls(self):
        """添加自定义URL用于导入页面"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.import_excel_view, name='englishword_import_excel'),
        ]
        return custom_urls + urls
    
    def media_count(self, obj):
        return obj.media_files.count()
    media_count.short_description = "Media Count"
    
    def save_model(self, request, obj, form, change):
        # 如果是新创建的单词，自动设置创建者为当前用户
        if not obj.pk:
            obj.creator = request.user
        super().save_model(request, obj, form, change)
    
    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related('creator').prefetch_related('tags')
        # 如果是超级用户，显示所有单词
        if request.user.is_superuser:
            return qs
        # 否则只显示当前用户创建的单词
        return qs.filter(creator=request.user)
    
    def has_change_permission(self, request, obj=None):
        # 如果是超级用户，允许编辑所有单词
        if request.user.is_superuser:
            return True
        # 如果obj存在，只允许创建者编辑
        if obj is not None and obj.creator != request.user:
            return False
        return True
    
    def has_delete_permission(self, request, obj=None):
        # 同样应用编辑权限规则到删除权限
        return self.has_change_permission(request, obj)
    
    # <font color="red">**新增点：重写表单字段的queryset，限制标签选择范围**</font>
    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "tags":
            # 根据用户权限过滤可选的标签
            if request.user.is_superuser:
                # 管理员可以看到所有标签
                kwargs["queryset"] = models.Tag.objects.all()
            else:
                # 普通用户只能看到自己创建的标签
                kwargs["queryset"] = models.Tag.objects.filter(creator=request.user)
        return super().formfield_for_manytomany(db_field, request, **kwargs)
    
    # 导出为Excel功能
    def export_as_excel(self, request, queryset):
        """导出选中的单词为Excel"""
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Words"
        
        # <font color="red">**修改点：在导出中添加标签列**</font>
        headers = ['单词', '解释', '备注', '标签', '创建者', '创建时间']
        ws.append(headers)
        
        # 添加数据
        for word in queryset:
            ws.append([
                word.title, 
                word.explanation, 
                word.notes or '',
                word.get_tag_names(),  # <font color="red">**新增标签数据**</font>
                word.creator.username,
                word.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="english_words.xlsx"'
        
        # 保存工作簿到响应
        wb.save(response)
        
        self.message_user(request, f"成功导出 {queryset.count()} 个单词到Excel")
        return response
    export_as_excel.short_description = "导出选中的单词为Excel"
    
    # 导出为CSV功能
    def export_as_csv(self, request, queryset):
        """导出选中的单词为CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="english_words.csv"'
        
        writer = csv.writer(response)
        # <font color="red">**修改点：在导出中添加标签列**</font>
        writer.writerow(['单词', '解释', '备注', '标签', '创建者', '创建时间'])
        
        for word in queryset:
            writer.writerow([
                word.title, 
                word.explanation, 
                word.notes or '',
                word.get_tag_names(),  # <font color="red">**新增标签数据**</font>
                word.creator.username,
                word.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        self.message_user(request, f"成功导出 {queryset.count()} 个单词到CSV")
        return response
    export_as_csv.short_description = "导出选中的单词为CSV"
    
    # 导出为JSON功能
    def export_as_json(self, request, queryset):
        """导出选中的单词为JSON"""
        data = []
        for word in queryset:
            data.append({
                'title': word.title,
                'explanation': word.explanation,
                'notes': word.notes,
                # <font color="red">**修改点：在导出中添加标签数据**</font>
                'tags': [tag.name for tag in word.tags.all()],
                'creator': word.creator.username,
                'created_at': word.created_at.isoformat(),
                'media_files': [media.file.url for media in word.media_files.all()]
            })
        
        response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="english_words.json"'
        
        self.message_user(request, f"成功导出 {queryset.count()} 个单词到JSON")
        return response
    export_as_json.short_description = "导出选中的单词为JSON"
    
    # Excel导入视图
    def import_excel_view(self, request):
        """处理Excel文件导入的视图"""
        context = {
            'title': 'impoert English Words from Excel',
            **self.admin_site.each_context(request),
        }
        
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            
            try:
                # 读取Excel文件
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # 使用事务确保数据一致性
                with transaction.atomic():
                    # 从第二行开始读取（跳过标题行）
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        # 跳过空行
                        if not row or not row[0]:
                            continue
                            
                        try:
                            # 解析行数据
                            title = str(row[0]).strip() if row[0] else ''
                            explanation = str(row[1]).strip() if row[1] else ''
                            notes = str(row[2]).strip() if row[2] else ''
                            # <font color="red">**新增点：解析标签数据（第4列）**</font>
                            tags_str = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                            
                            # 验证必要字段
                            if not title:
                                raise ValueError("单词标题不能为空")
                            if not explanation:
                                raise ValueError("单词解释不能为空")
                            
                            # 创建单词对象
                            word = models.EnglishWord(
                                title=title,
                                explanation=explanation,
                                notes=notes,
                                creator=request.user  # 自动设置当前用户为创建者
                            )
                            word.full_clean()  # 验证数据
                            word.save()
                            
                            # <font color="red">**新增点：处理标签关联**</font>
                            if tags_str:
                                tag_names = [tag_name.strip() for tag_name in tags_str.split(',') if tag_name.strip()]
                                for tag_name in tag_names:
                                    # 获取或创建标签（仅限当前用户创建的标签）
                                    tag, created = models.Tag.objects.get_or_create(
                                        name=tag_name,
                                        creator=request.user,
                                        defaults={'name': tag_name, 'creator': request.user}
                                    )
                                    word.tags.add(tag)
                            
                            success_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            errors.append(f"第{row_num}行错误: {str(e)}")
                
                # 显示结果消息
                if success_count > 0:
                    messages.success(request, f"成功导入 {success_count} 个单词")
                if error_count > 0:
                    messages.warning(request, f"有 {error_count} 个单词导入失败")
                    # 只显示前5个错误，避免消息过长
                    for error in errors[:5]:
                        messages.error(request, error)
                    if len(errors) > 5:
                        messages.info(request, f"... 还有 {len(errors) - 5} 个错误未显示")
                        
            except Exception as e:
                messages.error(request, f"文件处理错误: {str(e)}")
        
        return render(request, 'admin/englishword_import_excel.html', context)

@admin.register(models.Tag)
class TagAdmin(admin.ModelAdmin):
    """标签管理"""
    list_display = ['name', 'creator', 'created_at', 'word_count']
    list_filter = ['created_at']
    search_fields = ['name', 'creator__username']
    actions = None  # 禁用批量操作
    
    # 排除创建者字段，使其不在表单中显示
    exclude = ['creator']
    
    def word_count(self, obj):
        return obj.english_words.count()
    word_count.short_description = 'Associated Word Count'
    
    def save_model(self, request, obj, form, change):
        # 如果是新创建的标签，自动设置创建者为当前用户
        if not obj.pk:
            obj.creator = request.user
        super().save_model(request, obj, form, change)
    
    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related('creator')
        # 如果是超级用户，显示所有标签
        if request.user.is_superuser:
            return qs
        # 否则只显示当前用户创建的标签
        return qs.filter(creator=request.user)
    
    def has_change_permission(self, request, obj=None):
        # 如果是超级用户，允许编辑所有标签
        if request.user.is_superuser:
            return True
        # 如果obj存在，只允许创建者编辑
        if obj is not None and obj.creator != request.user:
            return False
        return True
    
    def has_delete_permission(self, request, obj=None):
        # 同样应用编辑权限规则到删除权限
        return self.has_change_permission(request, obj)



# 先去掉了这部分的展示
#@admin.register(models.EnglishWordMedia)
class EnglishWordMediaAdmin(admin.ModelAdmin):
    """英文单词多媒体文件管理"""
    list_display = ['word', 'file', 'uploaded_at']
    list_filter = ['uploaded_at']
    search_fields = ['word__title', 'file']
    actions = None  # 禁用批量操作

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # 如果是超级用户，显示所有文件
        if request.user.is_superuser:
            return qs
        # 否则只显示当前用户创建的文件（通过关联的单词的创建者来过滤）
        return qs.filter(word__creator=request.user)

    def has_delete_permission(self, request, obj=None):
        # 超级用户拥有删除权限
        if request.user.is_superuser:
            return True
        # 如果obj存在，检查其关联的单词是否是当前用户创建的
        if obj is not None and obj.word.creator != request.user:
            return False
        return True

    def has_change_permission(self, request, obj=None):
        return self.has_delete_permission(request, obj)
    
    def get_readonly_fields(self, request, obj=None):
        # 在编辑页面中，将word字段设置为只读
        if obj:  # obj存在表示正在编辑现有对象
            return ['word'] + list(super().get_readonly_fields(request, obj))
        return super().get_readonly_fields(request, obj)

















